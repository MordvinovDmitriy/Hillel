#!/usr/bin/python3

import csv
import openpyxl

wb = openpyxl.Workbook()
csv_sheet = wb.active
csv_sheet.title = "CSV"

with open('json2csv.csv', 'r') as file_csv:
    csv_reader = csv.reader(file_csv, delimiter=';')
    for col_index, row in enumerate(csv_reader):
        for row_index, value in enumerate(row):
            # cell = csv_sheet.cell(row=col_index + 1, column=row_index + 1)
            cell = csv_sheet.cell(row=row_index + 1, column=col_index + 1)
            if value == 'ID':
                value = value.lower()
            elif value == 'Ім\'я':
                value = 'name'
            elif value == 'Телефон':
                value = 'phone'
            cell.value = value
max_col = csv_sheet.max_column
csv_sheet.insert_rows(idx=0, amount=1)
for i in range(1, max_col):
    cell = csv_sheet.cell(row=1, column=i+1)
    cell.value = f'Person {i}'
for col in csv_sheet.columns:
    column = col[0].column_letter
    if column != 'A':
        csv_sheet.column_dimensions[column].width = 15
csv_sheet.delete_rows(4, 1)
wb.save('My_wb.xlsx')
